from fastapi import UploadFile
from openpyxl import load_workbook

from io import BytesIO

from ..models.file_content import FileContent, FileData

def convert_json(excel_file: UploadFile) -> FileContent:
  # excel_data = {
  #   'FileName': excel_file.filename,
  #   'Size': excel_file.size,
  #   'RowCount': 0,
  #   'Data': []
  # }
  excel_data = FileContent(FieldName=excel_file.filename,
                           Size=excel_file.size)
  
  with BytesIO(excel_file.file.read()) as file_content:
    workbook = load_workbook(file_content, read_only=True)
    
    total_rows = 0
    for sheet in workbook.worksheets:
      # sheet_data = {
      #   'SheetName': sheet.title,
      #   'RowCount': sheet.max_row - 1,
      #   'Rows': []
      # }
      sheet_data = FileData(SheetName=sheet.title,
                            RowCount=sheet.max_row-1)

      total_rows += sheet.max_row - 1
      header_row = [cell.value.replace('_', '') for cell in sheet[1]]
      for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = [str(cell).strip() for cell in row]
        sheet_data.rows.append(dict(zip(header_row, row_data)))
      excel_data.file_data.append(sheet_data)
      
    excel_data.row_count = total_rows
  return excel_data